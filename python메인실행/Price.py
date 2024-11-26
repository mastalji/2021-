#pip install pandas-datareader,yfinance,matplotlib,xlsxwriter,openpyxl
        
from pandas_datareader import data as pdr
from keras.models import Sequential
from keras.layers import Flatten, Dense, Input

import yfinance as yf
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl
import sys            #오류 발생시 강제종류 함수 

print("Wait",end="\n")
yf.pdr_override()    #데이터 읽는 방식 크롤링으로 변경-주가 데이터 불러올 수 있게됨

#코스피 코스닥 가져오기

url = 'https://kind.krx.co.kr/corpgeneral/corpList.do'
kospi_code = pd.read_html(url+"?method=download&marketType=stockMkt")[0]
kosdaq_code = pd.read_html(url+"?method=download&marketType=kosdaqMkt")[0]

kospi = kospi_code[['회사명','종목코드']]                             #행은 데이터, 열은 회사명 종목코드
kosdaq = kosdaq_code[['회사명','종목코드']]                           #행은 데이터, 열은 회사명 종목코드 

print("Running",end='\n')

def make_code_ks(x):
    x = str(x)
    return '0'*(6-len(x))+x+'.KS'

def make_code_kq(x):
    x = str(x)
    return '0'*(6-len(x))+x+'.KQ'   

kospi['종목코드'] = kospi['종목코드'].apply(make_code_ks)
kosdaq['종목코드'] = kosdaq['종목코드'].apply(make_code_kq)


#엑셀 파일 생성

kospi=kospi['종목코드'].to_list()
kosdaq=kosdaq['종목코드'].to_list()
workbook = xlsxwriter.Workbook('test.xlsx')
worksheet = workbook.add_worksheet()

cell_format = workbook.add_format({'border': 1})
worksheet.set_column('A:A', 10)
worksheet.set_column('B:B', 10)

for i in range(len(kospi)):
    worksheet.write('A%d'%(i+1),'%s'%kospi[i])
for i in range(len(kosdaq)):    
    worksheet.write('B%d'%(i+1),'%s'%kosdaq[i])
    
workbook.close()

#이렇게 하면 test.xlsx라는 파일이 생성되거나 수정되서 각 회사의 종목이름이됨.


#주가 함수

class finance:
    
    def __init__(self):
        self.s=0
        self.stock = 0
        self.stockmoney = 0
        
    def mainfunc(self,num):
        self.main=pdr.get_data_yahoo('%s'%num, start='2020-01-01', end='2021-12-31')        #시작날짜 마지막 날짜 수동입력 해야함.

#주식데이터 엑셀 파일에서 열어오기

jusik=[]                                                       #모든 주식의 모든 정보를 저장할 리스트. 이 프로그램에서 가장 중요한 리스트

flname=input("Enter file name: ")                              #여따 걍 test.xlsx치삼
print()
    
wb = openpyxl.load_workbook(filename='%s'%flname)              #test.xlsx
ws = wb.active                                                 #ws=worksheet
stockindex = 0                                                 #다룰 주식 총 개수(아마 3800개정도)                                                                                                                           

a=input("Kospi or Kosdaq: ")                                   #코스피 아니면 코스닥
print()

if a=="Kospi":
    a='A'
elif a=="Kosdaq":
    a='B'

b=input("How many company: ")
print()

for cell in ws[a]:                                             #cell은 엑셀시트의 첫번째 칸 

    sam = finance()
    sam.mainfunc(cell.value)                                   #cell.value=종목코드
    jusik.append(sam.main)                                     #listen[j]=samsung     samsungopen=list(samsung.Open)=list(listen[j].Open)
    stockindex+=1
    if stockindex==int(b):
        break  


#5일선 10일선 20일선


for s in range(stockindex):
    moveavg5 = []
    moveavg10 = []
    moveavg20 = []
    
    for k in range(len(jusik[s])):      
        five=0
        ten=0
        twenty=0
        open=list(jusik[s]['Close'])
              
        if k>=5:
            for i in range(k-5,k):
                five+=open[i]
        if k>=10:
            for i in range(k-10,k):
                ten+=open[i]
        if k>=20:
            for i in range(k-20,k):
                twenty+=open[i]
                
        moveavg5.append(five/5) 
        moveavg10.append(ten/10)
        moveavg20.append(twenty/20)

    jusik[s]['moveavg5']=list(moveavg5)
    jusik[s]['moveavg10']=list(moveavg10)
    jusik[s]['moveavg20']=list(moveavg20)


#추세선 그려서 기울기랑 R값 구하기

X=[]
y=[]

for s in range(stockindex):
    a=int(len(jusik[s]['Close'])/2)    
    x=np.arange(a)
    Y=[]
    for k in range(a):                           #이렇게 하면 전체범위의 반만 주가가 생김
        Y.append(jusik[s]['Close'][k])
    fit_line = np.polyfit(x,Y,1)
    X.append(fit_line[0])
    y.append(jusik[s]['Close'][len(jusik[s])-1]/jusik[s]['Close'][int(len(jusik[s])/2)])




# 군집화

df = pd.DataFrame(data = list(zip(X,y)), columns = ['chu','expect'])
df.head(5)

q3 = df.quantile(0.75) 
q1 = df.quantile(0.25)
iqr = q3 - q1
print(q3)
print(q1)
print(iqr)

# '추세선' 열에 대하여 이상치 여부를 판별해주는 함수
def is_chu_outlier(df):
    chu_score = df['chu']
    if chu_score > q3['chu'] + 1.5 * iqr['chu'] or chu_score < q1['chu'] - 1.5 * iqr['chu']:
        return True
    else:
        return False
# apply 함수를 통하여 각 값의 이상치 여부를 찾고 새로운 열에 결과 저장
df['chu_이상치여부'] = df.apply(is_chu_outlier, axis = 1) # axis = 1 지정 필수


# 'expect' 열에 대하여 이상치 여부를 판별해주는 함수
def is_expect_outlier(df):
    expect_score = df['expect']
    if expect_score > q3['expect'] + 1.5 * iqr['expect'] or expect_score < q1['expect'] - 1.5 * iqr['expect']:
        return True
    else:
        return False

# apply 함수를 통하여 각 값의 이상치 여부를 찾고 새로운 열에 결과 저장
df['expect_이상치여부'] = df.apply(is_expect_outlier, axis = 1) # axis = 1 지정 필수


# 이상치인 행은 제거하여 필터링
# 이상치여부를 나타내는 열 제거

df

df_trim = df.loc[df['chu_이상치여부'] == False]
del df_trim['chu_이상치여부']
df_trim = df.loc[df['expect_이상치여부'] == False]
del df_trim['expect_이상치여부']
df_trim




#인공지능 학습

model1 = Sequential()
model1.add(Dense(6,input_dim=1,activation='ELU'))
model1.add(Dense(24))
model1.add(Dense(1000))
model1.add(Dense(100))
model1.add(Dense(1))

model1.compile(loss='mean_squared_error',
               optimizer='adam')

model1.fit(np.array(X)-min(X),np.array(y)-min(y),epochs=20)




#계산된 값에 c/R^2을 뺀게 그 주식에 대한 평가율.
#당장은 입력한 주식에 대하여 평가 진행

K=[]
for s in range(stockindex):
    J=[]
    l=np.arange(len(jusik[s]['Close']))
    for k in range(len(jusik[s]['Close'])):
        J.append(jusik[s]['Close'][k])
    fit_line = np.polyfit(l,J,1)
    K.append(fit_line[0])
    
Process=np.array(model1.predict(np.array(K)-0.5)*100)

print(Process)














