
#pip install pandas-datareader,yfinance,matplotlib,xlsxwriter,openpyxl
        
print()
print("Wait second........")
print()

from pandas_datareader import data as pdr
import yfinance as yf
import matplotlib.pyplot as plt
import xlsxwriter
import openpyxl
import sys      #if unexpectable error

yf.pdr_override()    #데이터 읽는 방식 크롤링으로 변경-주가 데이터 불러올 수 있게됨


#주가 함수

class finance:
    
    def __init__(self):
        self.s=0
        self.stock = 0
        self.stockmoney = 0
        
    def mainfunc(self,num):
        self.main=pdr.get_data_yahoo('%s'%num, start='2021-01-01', end='2022-11-01') #005930.KS 005380.KS 0005.HK   시작날짜 마지막 날짜

#리스트 이름 정의

sat=[]  #주가        satshareprice
sbt=[]  #주수        sharenumber
sct=[]  #0,1저
mesu=0.5  #매수비율 
medo=1  #매도비율 ::그냥 매도는 전액매도로 하자.
maxx=[]  #종목 max
k=1
j=0

#kama=  #5-20, 엘리어트 간 비율
anna=[]
banna=[]

#주식종목 읽기

nancy=input("Choose data input type: by excel = press'E' | by handwork = press'H' ")
print()
listen=[]
n=0

if nancy!='E'and nancy!="H":
    
    print("UNEXPECTED ERROR !")
    sys.exit()

if nancy=='E':

    print("Available File: stocksave.xlsx, activestocksave.xlsx, activestocksave2.xlsx, downstocksave.xlsx, downstocksave2.xlsx, downstocksave3.xlsx, samsung.xlsx, shortstocksave.xlsx, shortstocksave2.xlsx, shortstocksave3.xlsx, upstocksave.xlsx, upstocksave2.xlsx ")
    print()
    st=input("Enter file name: ")
    print()
    
    wb = openpyxl.load_workbook(filename='%s'%st) #stocksave.xlsx
    ws = wb.active
    stockindex = 0
    k=0
    
    for cell in ws['A']:

        stockindex+=1
        
    for i in range(stockindex):

        listen.append(chr(k))
        k+=1

    stockindex=0
    
    for cell in ws['A']:

        sam = finance()
        print('stock: %s'%cell.value)
        b= cell.value
        sam.mainfunc(b)
        print(sam.main)
        print(sam.main.index)
        print(sam.main.columns)          #date
        listen[stockindex] = sam.main    #listen[j]=samsung     samsungopen=list(samsung.Open)=list(listen[j].Open)
        stockindex+=1

        for i in range(2):
            print("*")

        print("Perfect!")

        for i in range(2):
            print("*")
        n=stockindex

    for ak in range(n):

        sat.append(0)
        sbt.append(0)
        sct.append(0)
        
#money, stock

money=100000000 #1억

#output

if nancy=='H':

    n=int(input("How many stocks do you want to manage "))
    print()

    for i in range(n):

        sat.append(0)
        sbt.append(0)
        sct.append(0)

    for i in range(n):

        listen.append(chr(k))
        k+=1
    
    for sam in listen:

        sam=finance()
        print("Please enter the serial numbers of each stock ",end='')
        if n!=1:
            print()
        b=input()
        print()
        print("Wait second.......")
        sam.mainfunc(b)
        print(sam.main)
        print(sam.main.index)
        print(sam.main.columns)  #date
        listen[j] = sam.main    #listen[j]=samsung  samsungopen=list(samsung.Open)=list(listen[j].Open)
        j+=1

        for i in range(3):
            print("*")
        print("Perfect!")
        for i in range(3):
            print("*")

#이동평균선

#5일선
            
kno=input("Show movingaverage lines? Press 'Y'|'N' ")
print()

for kl in range(n):
    moveavg5 = []
    moveavg20 = []
    open=list(listen[kl].Open)
    for i in range(4):
        moveavg5.append(open[i])
    for i in range(4,len(open)):
        avg5 = 0
        a = 0

        for k in range(0,5):
            a += open[i-k]

        avg5 = a/5
        moveavg5.append(avg5)

    if kno=='Y':
        print(moveavg5)
        
#20일선

    for i in range(19):
        moveavg20.append(open[i])

    for i in range(19,len(open)):
        avg20 = 0
        a = 0

        for k in range(0,20):
            a += open[i-k]

        avg20 = a/20
        moveavg20.append(avg20)
        
    if kno=='Y':
        print(moveavg20)
        
    plt.plot(listen[kl].index, listen[kl].Open, 'b', label='Samsung Electronics')
    plt.plot(list(listen[kl].index), moveavg5, 'r', label='Samsung Electronics')
    plt.plot(list(listen[kl].index), moveavg20, 'y', label='Samsung Electronics')

#buy algorithm
 
bhc=len(list(listen[0].Open))   #길이 초기화

for kama in range(n):

    if bhc>len(list(listen[kama].Open)):
        bhc=len(list(listen[kama].Open))


for h in range(bhc+1):
    maxx.append(0)
        
k=0
kkki=1
kkkki=0
print("investment=100000000")
print()

for i in range(21,bhc):
     
    s=0
    simbol=0
    
    for bici in range(n):   #open도 변수

        sct[bici]=0
        open=list(listen[bici].Open)
        sat[bici]=open[i]   

        if k==0:

            sbt[bici]=int(money/n/sat[bici])
            money-=sat[bici]*sbt[bici]
            k=1
#a.
            
#b. 이동평균선 알고리즘

    #1. 평균선
            
        moveavg5=0
        for em in range(5): 
            moveavg5+=open[i-em]/5

        moveavg4=0
        for em in range(5):
            moveavg4+=open[i-em-1]/5

        moveavg20=0
        for em in range(20):
            moveavg20+=open[i-em]/20

        if open[i]>maxx[i-22]:
            maxx[i-21]=open[i]
            
    #2. 매수조건
            if moveavg5 > moveavg4 and moveavg5 and moveavg5 > moveavg20: 
                sct[bici]=1
        else:
            maxx[i-21]=maxx[i-22]

    #3. 매도조건
            
        if moveavg5 <= moveavg20 and (maxx[i-21]*70+(kkkki*30))/100 > open[i]:
            sct[bici]=0
            money+=open[i]*int(sbt[bici]*medo) #매도
            sbt[bici]-=int(sbt[bici]*medo)
            kkki=0
            kkkki=open[i]   
            
    #4. 전부팔기
            
        if i==bhc-1:
            for impact in range(0,n):
                money+=sat[impact]*sbt[impact]
                sbt[impact]=0
            break
         
    if i!=bhc-1:

        for a in range(n):
            s+=sct[a]

        for j in range(n):

            if s==0:
                continue

            if s!=0:
                kam=sct[j]*money*mesu/n #매수  
                sbt[j]+=int(kam/(sat[j]*s))
                money-=sat[j]*int(kam/(sat[j]*s))
                
    print("proceeds=%d number of shares (nos)"%money,end=' ')

    for o in range(0,n):
        print("nos%d="%(o+1),"%d"%sbt[o],end=' ')
    
    print()
    print()
    kkki+=1

print()
print("investment=100000000",end='  ')
print("totalequal=%d"%money, end=' ')
print(" ascentrate=%f"%(money/100000000)," %d%%"%(money/1000000))
print()
print("Wow! Delicous!")
print()

#plot

plt.show()

'''#exel
    workbook = xlsxwriter.Workbook('stock.xlsx')
    worksheet = workbook.add_worksheet()

    cell_format = workbook.add_format({'border': 1})
    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('C:C', 10)

    for i in range(len(sam.main.Open)):
        worksheet.write('A%d'%(i+1),'%d'%sam.main.Open[i])
        worksheet.write('B%d'%(i+1),'%d'%moveavg5[i])
        worksheet.write('C%d'%(i+1),'%d'%moveavg20[i])
    
    workbook.close()
'''
