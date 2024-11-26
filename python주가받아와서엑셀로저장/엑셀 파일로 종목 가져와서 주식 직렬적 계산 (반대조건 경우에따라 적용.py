#cmd
#pip install pandas,yfinace,matplotlib,xlsxwriter

from pandas_datareader import data as pdf
import yfinance as yf
import matplotlib.pyplot as plt
import xlsxwriter
import openpyxl
import random

yf.pdr_override()

#총 통장
global totmoney
totmoney = 0
global totdiv
totdiv = 0
global totreverse
totreverse = 0
global dontbuy
dontbuy = 0

#stock class
class finance:
    def __init__(self):
        self.stock = 0      #주식 주 수 
        self.stockmoney = 100000000 #1억 주식 통장
    def mainfunc(self,stockname):
        self.main=pdf.get_data_yahoo('%s'%stockname, start='2020-10-26', end='2021-10-26')

#주식종목 읽기
wb = openpyxl.load_workbook(filename='stocksave.xlsx')
ws = wb.active
stockindex = 0

for cell in ws['A']:
    sam = finance()
    print('stock: %s'%cell.value)
    stockindex += 1
    b= cell.value
    sam.mainfunc(b)
#    print(sam.main)
#    print(sam.main.index)
#    print(sam.main.columns)  #date

#for가 계속 이어짐

    def visual(x):
            return x*1000
#봉차트 (close, open)
    close1 = list(sam.main.Close)

#이동평균선

    #5일선
    open1 = list(sam.main.Open)
    moveavg5 = []
    for i in range(4):
        moveavg5.append(0)
    for i in range(4,len(open1)):
        avg5 = 0
        a = 0
        for k in range(0,5):
            a += open1[i-k]
        avg5 = a/5
        moveavg5.append(avg5)
#    print(moveavg5)


    #20일선
    moveavg20 = []
    for i in range(19):
        moveavg20.append(0)
    for i in range(19,len(open1)):
        avg20 = 0
        a = 0
        for k in range(0,20):
            a += open1[i-k]
        avg20 = a/20
        moveavg20.append(avg20)
#    print(moveavg20)

    #유전 알고리즘 테스트 반복
    returnindex = 1
    reverse = 0
    while True:
        sam.stockmoney = 100000000
        sam.stock = 0
        #유전적 알고리즘 적용용 변수 설정
        div = 1
        lastdiv = 1
        maxmoney = 100000000
        def sumdev(k): #돈 넣는 비율 계산
            return k*div
        
#5,20이동평균선 알고리즘
        dayindex = list(sam.main.index)
        buyindex = {}
        for i in dayindex:
            buyindex[i] = 0
    
        sellindex = {}
        for i in dayindex:
            sellindex[i] = 0.5
    
        #1. 매수조건 (첫날에는 구매하지 않음)
        buying = 0
        for i in range(21,len(open1)):
            if open1[i] < 1:
                continue
            if reverse == 0:#정상 조건
                if open1[i] > open1[i-1] and open1[i] > moveavg5[i] and moveavg5[i] > moveavg20[i] and close1[i-1] < open1[i] and sam.stockmoney > open1[i]:
                    buyindex[dayindex[i]] = 1
                    buying = int(sumdev(int(sam.stockmoney/int(open1[i]))))
                    sam.stock += buying
                    sam.stockmoney -= buying*int(open1[i])
            
        #2. 매도조건
                if moveavg5[i] <= moveavg20[i] and sam.stock > 0:
                    sellindex[dayindex[i]] = 2
                    selling = int((sam.stock))
                    sam.stockmoney += selling*int(open1[i])
                    sam.stock -= selling
                    
            if reverse == 1:#반대조건
                if moveavg5[i] <= moveavg20[i] and sam.stockmoney > open1[i]:
                    buyindex[dayindex[i]] = 1
                    buying = int(sumdev(int(sam.stockmoney/int(open1[i]))))
                    sam.stock += buying
                    sam.stockmoney -= buying*int(open1[i])
            
        #반대매도조건
                if open1[i] > open1[i-1] and open1[i] > moveavg5[i] and moveavg5[i] > moveavg20[i] and close1[i-1] < open1[i] and sam.stock > 0:
                    sellindex[dayindex[i]] = 2
                    selling = int((sam.stock))
                    sam.stockmoney += selling*int(open1[i])
                    sam.stock -= selling 
                    

    #3. 결과
        sam.stockmoney += sam.stock*int(open1[-1])
        if sam.stockmoney < 100000000:
            reverse = 1
#        print("주식 종목:%s"%b,"결과:%d"%sam.stockmoney)
#        print(div)
#        print(".")
#        print(returnindex)
        maxmoney = max(sam.stockmoney,maxmoney)

    #4. 유전 알고리즘 반복 break
    #   if money/100000000 >= 1.1: 
    #       print(returnindex)
    #       break
        if returnindex == 1000:
            if maxmoney != 100000000:
                print(returnindex)
                print(maxmoney)
                print(lastdiv)
                print(maxmoney/100000000)
                if lastdiv != 1:
                    a = input("div 1아님")
                if reverse == 1:
                    totreverse += 1
                    pass
#                   a = input("reversed")
                totmoney += maxmoney
                totdiv += lastdiv
                break
            else:
                print(returnindex)
                print(sam.stockmoney)
                print(div)
                print(sam.stockmoney/100000000)
                print("사지 마세요")
                dontbuy += 1
                totmoney += sam.stockmoney
                totdiv += div
                break
        
    #5. break 실패시 다른 div 정하기
        returnindex += 1
        if maxmoney == sam.stockmoney:
            lastdiv = div
            up = lastdiv*1.2
            down = lastdiv*0.8
            if up > 1:
                up = 1
            div = random.uniform(down,up)
        else:
            div = lastdiv
            up = lastdiv*1.2
            down = lastdiv*0.8
            if up > 1:
                up = 1
            div = random.uniform(down,up)
        
#exel
    '''
    workbook = xlsxwriter.Workbook('stock%s.xlsx'%b)
    worksheet = workbook.add_worksheet()

    cell_format = workbook.add_format({'border': 1})
    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('C:C', 10)

    for i in range(len(sam.main.Open)):
        worksheet.write('A%d'%(i+1),'%d'%sam.main.Open[i])
        worksheet.write('B%d'%(i+1),'%d'%moveavg5[i])
        worksheet.write('C%d'%(i+1),'%d'%moveavg20[i])
    
    workbook.close()
    '''

'''#plot 한번에 생성이 가능해야함.
    plt.plot(dayindex, sam.main.Open, 'b', label = '%s'%b)
    plt.plot(dayindex, moveavg5, 'r')
    plt.plot(dayindex, moveavg20, 'y')
    plt.scatter(dayindex, list(map(visual,buyindex.values())), c='b', s=0.5)
    plt.scatter(dayindex, list(map(visual,sellindex.values())), c='r', s=0.5)
    plt.xlabel('days')
    plt.ylabel('price')
    plt.text(dayindex[1],10000,'buyindex', fontdict={'color':'blue'})
    plt.text(dayindex[1],15000,'sellindex', fontdict={'color':'red'})
    plt.title('%s'%b)

    plt.show() '''

#총 결과
print('총 주식 개수: %d'%stockindex)
print('avg div: %f'%(totdiv/(stockindex)))
print('reverse 비율: %f'%((totreverse/(stockindex))*100))
print('총 사지마세요: %f'%((dontbuy/(stockindex))*100))
print('수익률: %f'%(totmoney/(stockindex*100000000)))
